"""
data_parser.py — reads all 6 Excel files live.

Exact column layout (0-indexed):
  TY-A (Sheet1):
    Feb: CC[3,4] ML[5,6] WT[7,8] STQA[9,10] HPC[11] IOT[12] MP-II[13] att%[14]
    Mar: CC[17,18] ML[19,20] WT[21,22] STQA[23,24] HPC[25] IOT[26] MP-II[27] att%[28]

  TY-B (Sheet2):
    Feb: CC[3,4] ML[5,6] WT[7,8] Blockchain[9,10] HPC[11] SW[12] MP-II[13] att%=formula
    Mar: CC[18,19] ML[20,21] WT[22,23] Blockchain[24,25] HPC[26] SW[27] MP-II[28] att%=formula

  TY-C (Sheet1):
    Feb: CC[3,4] ML[5,6] WT[7,8] STQA[9,10] HPC[11] PM[12] MP-II[13] att%[14]
    Mar: CC[18,19] ML[20,21] WT[22,23] STQA[24,25] HPC[26] PM[27] MP-II[28] att%=formula[29]
"""
import os, datetime, math
import openpyxl

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')

# subject → (feb_th, feb_pr, mar_th, mar_pr)  — None if no Pr col
SUBJ_COLS_A = {
    'CC':    (3, 4,   17, 18),
    'ML':    (5, 6,   19, 20),
    'WT':    (7, 8,   21, 22),
    'STQA':  (9, 10,  23, 24),
    'HPC':   (11, None, 25, None),
    'IOT':   (12, None, 26, None),
    'MP-II': (13, None, 27, None),
}
SUBJ_COLS_B = {
    'CC':         (3, 4,   18, 19),
    'ML':         (5, 6,   20, 21),
    'WT':         (7, 8,   22, 23),
    'Blockchain': (9, 10,  24, 25),
    'HPC':        (11, None, 26, None),
    'SW':         (12, None, 27, None),
    'MP-II':      (13, None, 28, None),
}
SUBJ_COLS_C = {
    'CC':    (3, 4,   18, 19),
    'ML':    (5, 6,   20, 21),
    'WT':    (7, 8,   22, 23),
    'STQA':  (9, 10,  24, 25),
    'HPC':   (11, None, 26, None),
    'PM':    (12, None, 27, None),
    'MP-II': (13, None, 28, None),
}

def _pct(v):
    if isinstance(v, (int, float)) and 0 <= v <= 100:
        return round(float(v), 2)
    if isinstance(v, str):
        for p in v.replace('\n',' ').split():
            try:
                f = float(p)
                if 0 <= f <= 100: return round(f, 2)
            except Exception: pass
    return None

def _detect_months(rows):
    months, seen = [], set()
    for row in rows[:6]:
        for v in row:
            label = None
            if isinstance(v, datetime.datetime):
                label = v.strftime('%B %Y')
            elif isinstance(v, (int, float)) and 40000 < v < 55000:
                d = datetime.date(1899,12,30) + datetime.timedelta(days=int(v))
                label = d.strftime('%B %Y')
            if label and label not in seen:
                seen.add(label); months.append(label)
    return months or ['February 2026', 'March 2026']

def _get(row, col):
    if col is None: return 0
    return row[col] if (col < len(row) and isinstance(row[col], (int,float))) else 0

def _parse_att(rows, subj_cols_map, feb_att_col, mar_att_col, compute_mar=False):
    months = _detect_months(rows)
    feb_label = months[0]
    mar_label = months[1] if len(months) > 1 else 'March 2026'

    feb_data, mar_data = {}, {}
    total_rows = {}          # batch → total lec row
    current = 'Batch1'

    for row in rows:
        if not any(v is not None for v in row): continue

        if (row[0] is None and isinstance(row[1], str) and
                'No. of lectures' in str(row[2] or '')):
            b = str(row[1]).strip()
            if b:
                current = b
                feb_data.setdefault(current, [])
                mar_data.setdefault(current, [])
                total_rows[current] = row
            continue

        v0 = row[0]
        if not (isinstance(v0,(int,float)) and 100 <= v0 <= 999): continue

        roll = int(v0)
        grn  = str(row[1]).strip() if row[1] else ''
        name = str(row[2]).strip() if row[2] else ''
        feb_data.setdefault(current, [])
        mar_data.setdefault(current, [])
        tr = total_rows.get(current, row)

        # Overall att%
        feb_pct = _pct(row[feb_att_col]) if (not compute_mar and feb_att_col < len(row)) else None
        if feb_pct is None:
            t = sum(_get(tr,c) for c in [3,4,5,6,7,8,9,10,11,12,13] if isinstance(tr[c] if c<len(tr) else None,(int,float)) and 0 < (tr[c] if c<len(tr) else 0) <= 30)
            s = sum(_get(row,c) for c in [3,4,5,6,7,8,9,10,11,12,13])
            feb_pct = round(s/t*100, 2) if t else 0

        if compute_mar:
            mar_cols_range = list(subj_cols_map.values())
            mar_indices = []
            for v in mar_cols_range:
                if v[2] is not None: mar_indices.append(v[2])
                if v[3] is not None: mar_indices.append(v[3])
            t_mar = sum(_get(tr,c) for c in mar_indices if isinstance(tr[c] if c<len(tr) else None,(int,float)) and 0 < (tr[c] if c<len(tr) else 0) <= 30)
            s_mar = sum(_get(row,c) for c in mar_indices)
            mar_pct = round(s_mar/t_mar*100, 2) if t_mar else 0
        else:
            mar_pct = _pct(row[mar_att_col]) if (mar_att_col < len(row)) else None
            if mar_pct is None:
                mar_indices = [v[2] for v in subj_cols_map.values() if v[2]] + [v[3] for v in subj_cols_map.values() if v[3]]
                t2 = sum(_get(tr,c) for c in mar_indices if isinstance(tr[c] if c<len(tr) else None,(int,float)) and 0 < (tr[c] if c<len(tr) else 0) <= 30)
                s2 = sum(_get(row,c) for c in mar_indices)
                mar_pct = round(s2/t2*100, 2) if t2 else 0

        # Subject-wise
        subjs = {}
        for subj, (ft, fp, mt, mp) in subj_cols_map.items():
            fa = _get(row,ft) + _get(row,fp)
            ft_total = _get(tr,ft) + _get(tr,fp)
            ma = _get(row,mt) + _get(row,mp)
            mt_total = _get(tr,mt) + _get(tr,mp)
            subjs[subj] = {
                'feb_att': fa, 'feb_total': ft_total,
                'mar_att': ma, 'mar_total': mt_total,
            }

        entry = {'roll':roll,'grn':grn,'name':name,'attendance_pct':feb_pct,'subjects':subjs}
        entry_m = {'roll':roll,'grn':grn,'name':name,'attendance_pct':mar_pct,'subjects':subjs}
        feb_data[current].append(entry)
        mar_data[current].append(entry_m)

    return {feb_label: feb_data, mar_label: mar_data}

def load_attendance_a():
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR,'TY-A_Attendances.xlsx'), read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    return _parse_att(rows, SUBJ_COLS_A, feb_att_col=14, mar_att_col=28)

def _split_batch(data, big_batch, new_b2, new_b3, split_at=25):
    """
    For each month in data, split big_batch into new_b2 (first split_at students)
    and new_b3 (the rest).  Modifies data in-place and returns it.
    """
    for month in data:
        students = data[month].pop(big_batch, [])
        data[month][new_b2] = students[:split_at]
        data[month][new_b3] = students[split_at:]
    return data

def load_attendance_b():
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR,'TYB_Attendance.xlsx'), read_only=True)
    ws = wb['Sheet2'] if 'Sheet2' in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    data = _parse_att(rows, SUBJ_COLS_B, feb_att_col=14, mar_att_col=29, compute_mar=True)
    # B1=25 (already correct), split B2(53) → B2(25) + B3(28)
    return _split_batch(data, 'B2', 'B2', 'B3', split_at=25)

def load_attendance_c():
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR,'TY-C_Attendance.xlsx'), read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    data = _parse_att(rows, SUBJ_COLS_C, feb_att_col=14, mar_att_col=29)
    # C1=25 (already correct), split C2(53) → C2(25) + C3(28)
    return _split_batch(data, 'C2', 'C2', 'C3', split_at=25)

# ── MARKS ──────────────────────────────────────────────────────────
def _clean(v):
    if isinstance(v,str) and v.strip().upper()=='AB': return 'AB'
    if isinstance(v,(int,float)): return int(v)
    return None

def _ise(c1,c2):
    a1,a2 = c1=='AB', c2=='AB'
    if a1 and a2: return 'AB'
    return (0 if (a1 or c1 is None) else int(c1)) + (0 if (a2 or c2 is None) else int(c2))

def _avg(i1,i2):
    v1 = 0 if i1=='AB' else (int(i1) if i1 is not None else 0)
    v2 = 0 if i2=='AB' else (int(i2) if i2 is not None else 0)
    return math.ceil((v1+v2)/2)

def _parse_marks(ws, subjects):
    rows = list(ws.iter_rows(values_only=True))
    students, start = [], 0
    for i,row in enumerate(rows):
        if isinstance(row[0],(int,float)) and 0 < row[0] <= 500:
            start=i; break
    for row in rows[start:]:
        if not (isinstance(row[0],(int,float)) and 0 < row[0] <= 500): continue
        roll=int(row[0]); grn=str(row[1]).strip() if row[1] else ''; name=str(row[2]).strip() if row[2] else ''
        marks={}; col=3
        for subj in subjects:
            if col+4>=len(row): break
            c1,c2=_clean(row[col]),_clean(row[col+1])
            c3,c4=_clean(row[col+3]),_clean(row[col+4])
            i1,i2=_ise(c1,c2),_ise(c3,c4)
            marks[subj]={'CO1':c1,'CO2':c2,'ISE1':i1,'CO3':c3,'CO4':c4,'ISE2':i2,'average':_avg(i1,i2)}
            col+=6
        students.append({'roll':roll,'grn':grn,'name':name,'subjects':marks})
    return students

def load_marks_a():
    wb=openpyxl.load_workbook(os.path.join(DATA_DIR,'TY-A_Marks.xlsx'),read_only=True)
    s=_parse_marks(wb.active,['CC','ML','HPC','STQA','IOT'])
    for x in s:
        if x['roll']<100: x['roll']+=100
    return s

def load_marks_b():
    wb=openpyxl.load_workbook(os.path.join(DATA_DIR,'TY-B_Marks.xlsx'),read_only=True)
    return _parse_marks(wb.active,['CC','ML','HPC','Blockchain','Semantic Web'])

def load_marks_c():
    wb=openpyxl.load_workbook(os.path.join(DATA_DIR,'TY-C_Marks.xlsx'),read_only=True)
    return _parse_marks(wb.active,['CC','ML','HPC','STQA','PM'])

DIVISION_SUBJECTS={'A':['CC','ML','HPC','STQA','IOT'],'B':['CC','ML','HPC','Blockchain','Semantic Web'],'C':['CC','ML','HPC','STQA','PM']}
DIVISION_ATT_SUBJECTS={'A':['CC','ML','WT','STQA','HPC','IOT','MP-II'],'B':['CC','ML','WT','Blockchain','HPC','SW','MP-II'],'C':['CC','ML','WT','STQA','HPC','PM','MP-II']}
DIVISION_PASSWORDS={'A':'divA@2024','B':'divB#2024','C':'divC$2024'}
LOADERS={'A':{'attendance':load_attendance_a,'marks':load_marks_a},'B':{'attendance':load_attendance_b,'marks':load_marks_b},'C':{'attendance':load_attendance_c,'marks':load_marks_c}}
